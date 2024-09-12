import os
import random
import json
import re
import hashlib
import time
from typing import List, Set, Dict, Tuple

from openpyxl import Workbook, load_workbook
from dataclasses import dataclass

# ======================== CONFIGURATION ========================
ROOT_DIRECTORY = 'traits'  # Location of your folder if not in the same directory
EXCEL_FILE = 'traits_info.xlsx'  # Name of the traits info excel spreadsheet


# ======================== TYPES ========================
@dataclass
class TraitInfo:
    inscription_id: str
    number: int
    type: str
    name: str
    weight: float
    blacklist: Set[int]
    whitelist: Set[int]

    def __hash__(self):
        return hash((self.type, self.name))


TraitsInfo = Dict[str, Dict[str, TraitInfo]]
TraitsMapping = Dict[int, TraitInfo]
FormattedInscription = List[Dict[str, str]]


# ======================== UTILITY FUNCTIONS ========================
def validate_inscription_id(inscription_id: str) -> bool:
    return re.match(r"^[\da-fA-F]{64}i\d+$", inscription_id) is not None


def get_positive_integer(message: str) -> int:
    while True:
        try:
            value = int(input(message))
            if value > 0:
                return value
            else:
                print("Please enter a positive number.")
        except ValueError:
            print("Invalid input. Please enter a valid number.")


# ======================== SPREADSHEET MANAGEMENT ========================
def create_spreadsheet(traits_dict: Dict[str, List[str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Traits Information"

    # Link to instructions, and merge cells with link in it
    sheet['A1'].value = 'Click here for detailed instructions on how to fill this out'
    sheet['A1'].hyperlink = 'https://github.com/Vanniix/generateorderv2/blob/main/README.md#how-to-use'
    sheet['A1'].style = 'Hyperlink'
    sheet.merge_cells('A1:C1')

    headers = [
        "Number", "Trait Type", "Trait Name", "Rarity (%)",
        "Blacklist", "Whitelist", "Inscription ID"
    ]
    sheet.append(headers)

    # Widen some columns to make it easier to work with
    widths = [8, 20, 20, 10, 15, 10, 40]
    for i, width in enumerate(widths):
        sheet.column_dimensions[chr(ord('A') + i)].width = width

    trait_number = 1
    for trait_type, traits in traits_dict.items():
        for trait in traits:
            row = [trait_number, trait_type, trait, '', '', '']
            sheet.append(row)
            trait_number += 1
        sheet.append([trait_number, trait_type, 'none', '', '', ''])
        trait_number += 1

    workbook.save(EXCEL_FILE)
    print(f"\nA file named '{EXCEL_FILE}' has been created. "
          "Please fill in the required information in this file before proceeding.")


def parse_int_set(comma_separated_list: str, error_name: str) -> Set[int]:
    number_list = set()
    str_list = [x.strip() for x in comma_separated_list.split(',') if len(x.strip()) > 0]

    for potential_number in str_list:
        try:
            float_value = float(potential_number)
            if not float_value.is_integer():
                raise ValueError
            number_list.add(int(float_value))
        except ValueError:
            raise Exception(
                f"{error_name} contains invalid entry '{potential_number}'. Only whole numbers are allowed.")
    return number_list


def convert_whitelist_to_blacklist(traits: TraitsInfo, trait_mapping: TraitsMapping) -> None:
    # A whitelist is the inverse of a blacklist, so for each whitelist,
    # we add to the blacklist everything but the whitelist
    for trait_group in traits.values():
        for trait in trait_group.values():
            whitelist = [trait_mapping[w] for w in trait.whitelist]
            while len(whitelist) > 0:
                # get other whitelisted traits in the same type
                trait_type = whitelist[0].type
                if trait_type == trait.type:
                    raise Exception(f'{trait.type}/{trait.name}: Cannot whitelist two traits in the same trait type, '
                                    'as traits in the same trait type never generate together')

                whitelisted_traits = set([t for t in whitelist if t.type == trait_type])
                equivalent_blacklist = [t.number for t in traits[trait_type].values() if t not in whitelisted_traits]
                trait.blacklist.update(equivalent_blacklist)
                whitelist = [x for x in whitelist if x not in whitelisted_traits]


def load_traits_info() -> Tuple[TraitsInfo, TraitsMapping]:
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active

    all_traits_info: TraitsInfo = {}
    trait_number_mapping: TraitsMapping = {}
    errors = []
    cumulative_weights: Dict[str, float] = {}

    # Determine the row with the headers (in case user deletes the top row that has the link)
    header_index = 1
    for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
        if row[0] == 'Number':
            header_index = row_index + 1
            break

    for row_index, row in enumerate(sheet.iter_rows(min_row=header_index+1, values_only=True), start=header_index+1):
        if not row or not any(row):
            continue

        trait_number, trait_type, trait_name, rarity, blacklist, whitelist, inscription_id = \
            (item if item is not None else '' for item in row)

        if inscription_id and not validate_inscription_id(inscription_id):
            errors.append(f"Row {row_index}: Invalid Inscription ID format.")
            continue

        weight = 1
        if rarity != '':
            try:
                weight = float(rarity)
                if weight < 0:
                    errors.append(f"Row {row_index}: Rarity must be greater than 0.")
                    continue
            except ValueError:
                errors.append(f"Row {row_index}: Rarity should be a number or left blank for equal rarities.")
                continue

        cumulative_weights[trait_type] = cumulative_weights.get(trait_type, 0) + weight

        blacklist_set, whitelist_set = set(), set()
        try:
            blacklist_set = parse_int_set(str(blacklist), 'Blacklist')
        except Exception as e:
            errors.append(f'Row {row_index}: {e}')

        try:
            whitelist_set = parse_int_set(str(whitelist), 'Whitelist')
        except Exception as e:
            errors.append(f'Row {row_index}: {e}')

        intersection = whitelist_set & blacklist_set
        if len(intersection) > 0:
            errors.append(
                f'Row {row_index}: These traits are both whitelisted and blacklisted: '
                f'{",".join([str(x) for x in intersection])}. You cannot whitelist and blacklist the same trait')

        if errors:
            continue

        if trait_type not in all_traits_info:
            all_traits_info[trait_type] = {}

        trait = TraitInfo(
            inscription_id=inscription_id,
            number=int(trait_number),
            type=trait_type,
            name=trait_name,
            weight=weight,
            blacklist=blacklist_set,
            whitelist=whitelist_set
        )
        all_traits_info[trait.type][trait.name] = trait
        trait_number_mapping[trait.number] = trait

    try:
        convert_whitelist_to_blacklist(all_traits_info, trait_number_mapping)
    except Exception as e:
        errors.append(str(e))

    if errors:
        print("\nErrors were found in the spreadsheet. Please review the messages below.")
        for error_message in errors:
            print(error_message)
        print("\nPlease correct the errors in the spreadsheet and try again.")
        exit()

    # Normalise weights so they add to 1
    for trait_type, total_weight in cumulative_weights.items():
        for trait_name, trait_info in all_traits_info[trait_type].items():
            trait_info.weight /= total_weight

    return all_traits_info, trait_number_mapping


# ======================== VALIDATION FUNCTIONS ========================
def validate_inscription_avoidance(inscription_collection: List[FormattedInscription], all_traits_info: TraitsInfo):
    inconsistencies = []

    for inscription_index, inscription in enumerate(inscription_collection, start=1):
        current_inscription_trait_numbers = set()

        for formatted_trait in inscription:
            trait = all_traits_info[formatted_trait["trait_type"]][formatted_trait["value"]]
            current_inscription_trait_numbers.add(trait.number)

        for formatted_trait in inscription:
            trait = all_traits_info[formatted_trait["trait_type"]][formatted_trait["value"]]
            conflicting_traits = current_inscription_trait_numbers & trait.blacklist

            if conflicting_traits:
                conflicts = [f"Trait #{num}" for num in conflicting_traits]
                inconsistency_info = {
                    "Inscription_number": inscription_index,
                    "Trait": f"{trait.type} - {trait.name}",
                    "Conflicts": conflicts
                }
                inconsistencies.append(inconsistency_info)

    return inconsistencies


def validate_traits(selected_traits: List[TraitInfo], trait_number_mapping: TraitsMapping):
    for trait in selected_traits:
        for blacklist_number in trait.blacklist:
            blacklist_trait = trait_number_mapping[blacklist_number]
            if blacklist_trait in selected_traits:
                return False

    return True


# ======================== METADATA GENERATION ========================
def generate_inscriptions(
        all_traits_info: TraitsInfo, trait_number_mapping: TraitsMapping, num_inscriptions: int
) -> Tuple[List[FormattedInscription], Dict, Dict[int, int]]:
    inscription_collection: List[FormattedInscription] = []
    traits_usage = {
        trait_type: {
            trait_name: {
                "count": 0,
                "rarity": f"{trait.weight * 100:.2f}"
            } for trait_name, trait in traits_info.items()
        } for trait_type, traits_info in all_traits_info.items()
    }
    generated_hashes = set()
    trait_count_distribution = {}

    # Provide user with update every second for long-running generations
    last_update_time = time.time()

    for _ in range(num_inscriptions):
        generation_attempts = 0
        while True:
            if time.time() > last_update_time + 1:
                last_update_time = time.time()
                print(f'Generated {len(inscription_collection)}/{num_inscriptions}')

            inscription_traits: List[TraitInfo] = []
            formatted_traits: FormattedInscription = []
            current_avoid_list: List[int] = []

            valid_combination = True
            trait_types = list(all_traits_info.values())
            random.shuffle(trait_types)
            for trait_group in trait_types:
                available_traits = [trait for trait in trait_group.values()
                                    if trait.number not in current_avoid_list
                                    and len(trait.blacklist & set([t.number for t in inscription_traits])) == 0]

                if len(available_traits) == 0:
                    valid_combination = False
                    break

                weights: List[float] = []

                for trait in available_traits:
                    # Dynamic weight calculation. The weights are recalculated each time based on the current trait
                    # usage, so if trait exclusions are skewing the trait distribution, the weights are calculated to
                    # correct for it
                    if trait.weight == 0:
                        weights.append(0)
                    else:
                        expected_number = num_inscriptions * trait.weight
                        # This is the number needed to be generated to reach the desired rarity. We add 1 to introduce
                        # a bit of extra entropy, otherwise it gets stuck generating the same set of traits towards
                        # the end
                        weights.append(max(0.0, expected_number - traits_usage[trait.type][trait.name]["count"]) + 1)

                if sum(weights) == 0:
                    valid_combination = False
                    break

                selected_trait: TraitInfo = random.choices(available_traits, weights=weights)[0]

                current_avoid_list.extend(selected_trait.blacklist)
                inscription_traits.append(selected_trait)
                if selected_trait.name != 'none':
                    formatted_traits.append({
                        "trait_type": selected_trait.type,
                        "value": selected_trait.name
                    })

            ordered_traits = [list(x.values())[0].type for x in list(all_traits_info.values())]
            inscription_traits.sort(key=lambda x: ordered_traits.index(x.type))
            formatted_traits.sort(key=lambda x: ordered_traits.index(x['trait_type']))
            string_repr = str([(trait.type, trait.name) for trait in inscription_traits])
            inscription_hash = hashlib.sha256(string_repr.encode()).hexdigest()
            if inscription_hash in generated_hashes or not valid_combination:
                generation_attempts += 1

                if generation_attempts >= 10000:
                    print(f"Unable to generate unique metadata after {generation_attempts} attempts. "
                          f"You may have exhausted all unique combinations. "
                          f"Please add more traits, or lower collection total size.")
                    return inscription_collection, traits_usage, trait_count_distribution

                continue

            generated_hashes.add(inscription_hash)
            if validate_traits(inscription_traits, trait_number_mapping):
                for trait in inscription_traits:
                    traits_usage[trait.type][trait.name]["count"] += 1

                num_traits = len(inscription_traits)

                trait_count_distribution[num_traits] = trait_count_distribution.get(num_traits, 0) + 1
                inscription_collection.append(formatted_traits)
                break

    # The dynamic weighting can result in some traits being more/less likely to occur towards the end. Shuffling will
    # make them uniform again
    random.shuffle(inscription_collection)

    traits_usage_statistics = {}
    for trait_type, traits in traits_usage.items():
        traits_usage_statistics[trait_type] = {}

        trait_type_count = sum(usage_info["count"] for usage_info in traits.values())

        for trait, usage_info in traits.items():
            trait_info = {
                "usage": f"{usage_info['count']} ({(usage_info['count'] / trait_type_count) * 100:.2f}%)",
                "rarity input": usage_info["rarity"]
            }

            if trait == 'none':
                trait_info["none_status"] = "Used" if usage_info['count'] > 0 else "Not used"

            traits_usage_statistics[trait_type][trait] = trait_info

    return inscription_collection, traits_usage_statistics, trait_count_distribution


# ======================== MAIN EXECUTION ========================
def main():
    print("Collection Metadata Generator\n")

    if not os.path.exists(EXCEL_FILE):
        trait_types = sorted([x for x in os.listdir(ROOT_DIRECTORY) if os.path.isdir(os.path.join(ROOT_DIRECTORY, x))],
                             key=lambda x: (int(x.split('.')[0]), x.split('.')[1]))
        traits_dict: Dict[str, List[str]] = {}

        for trait_type in trait_types:
            trait_type_name = trait_type.split('. ')[1]
            traits = sorted(os.listdir(os.path.join(ROOT_DIRECTORY, trait_type)))
            traits = [trait.split('.')[0] for trait in traits]
            traits_dict[trait_type_name] = [trait for trait in traits if len(trait) > 0]
        create_spreadsheet(traits_dict)

    input(f"\nPress Enter after you have updated the '{EXCEL_FILE}' file with the required information...")

    all_traits_info, trait_number_mapping = load_traits_info()
    num_inscriptions = get_positive_integer("\nEnter the number of Inscriptions you want to generate metadata for: ")

    inscription_collection, traits_usage_statistics, trait_count_distribution = \
        generate_inscriptions(all_traits_info, trait_number_mapping, num_inscriptions)

    inconsistencies = validate_inscription_avoidance(inscription_collection, all_traits_info)

    if inconsistencies:
        print("\nInconsistencies found in trait avoidance rules:")
        for inconsistency in inconsistencies:
            print(f"- {inconsistency}")

    with open('metadata.json', 'w') as file:
        json.dump([
            {
                "token id": f"#{i+1}",
                "attributes": item
            } for i, item in enumerate(inscription_collection)
        ], file, indent=4)

    traits_inscription_mapping = {}
    for trait_type, traits in all_traits_info.items():
        traits_inscription_mapping[trait_type] = {trait.name: trait.inscription_id for trait in traits.values()}

    with open('traits.json', 'w') as file:
        json.dump(traits_inscription_mapping, file, indent=4)

    # Prepare the summary for trait_usage_statistics.json
    summary = {
        "Total_inscriptions": num_inscriptions,
        "Trait_count_distribution": {f"{count}_traits": f"{amount} inscriptions"
                                     for count, amount in trait_count_distribution.items()},
        "Traits_usage": traits_usage_statistics
    }

    with open('trait_usage_statistics.json', 'w') as file:
        json.dump(summary, file, indent=4)

    print("\nInscription generation is complete. "
          "Check 'metadata.json', 'traits.json', and 'trait_usage_statistics.json' for the collection.")


if __name__ == '__main__':
    main()
